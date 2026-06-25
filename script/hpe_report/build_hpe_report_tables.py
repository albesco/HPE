#!/usr/bin/env python3
"""
Build HPE report tables from COCO-style GT JSON, YOLO/VitPose prediction JSON,
and validation metrics CSV files.

Default output:
  data/output/experiments/hpe_report/hpe_report_tables.xlsx

The script is intentionally headless/server-friendly: it does not require Excel,
LibreOffice, Codex, notebooks, or a display server.
"""

from __future__ import annotations

import argparse
import contextlib
import csv
import io
import json
import math
import os
import re
import sys
import tempfile
from dataclasses import dataclass
from pathlib import Path
from statistics import mean, median
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    from pycocotools.coco import COCO  # type: ignore
    from pycocotools.cocoeval import COCOeval  # type: ignore
except Exception:  # pragma: no cover - optional runtime dependency
    COCO = None
    COCOeval = None

COCO_KEYPOINTS = [
    "nose",
    "left_eye",
    "right_eye",
    "left_ear",
    "right_ear",
    "left_shoulder",
    "right_shoulder",
    "left_elbow",
    "right_elbow",
    "left_wrist",
    "right_wrist",
    "left_hip",
    "right_hip",
    "left_knee",
    "right_knee",
    "left_ankle",
    "right_ankle",
]

HEAD_KP = {"nose", "left_eye", "right_eye", "left_ear", "right_ear"}
BODY_KP_INDEXES = [i for i, kp in enumerate(COCO_KEYPOINTS) if kp not in HEAD_KP]
ALL_KP_INDEXES = list(range(len(COCO_KEYPOINTS)))

# Official COCO keypoint sigmas divided by 10, as used by pycocotools.
COCO_SIGMAS = np.array(
    [0.26, 0.25, 0.25, 0.35, 0.35, 0.79, 0.79, 0.72, 0.72,
     0.62, 0.62, 1.07, 1.07, 0.87, 0.87, 0.89, 0.89],
    dtype=np.float64,
) / 10.0

OKS_THRESHOLDS = np.round(np.arange(0.50, 0.96, 0.05), 2)


@dataclass(frozen=True)
class Prediction:
    image_id: int
    keypoints: List[Tuple[float, float, float]]
    score: float
    bbox: Optional[List[float]] = None


@dataclass(frozen=True)
class Scenario:
    scenario_id: str
    dataset_train: str
    dataset_test: str
    gt_key: str
    model_key: str
    model_label: str
    pred_key: str
    threshold: Optional[float]
    threshold_label: str


@dataclass
class EvalBundle:
    scenario: Scenario
    gt_path: Path
    pred_path: Path
    gt: Dict[str, Any]
    predictions: Dict[int, Prediction]
    error_rows: List[Dict[str, Any]]
    summary: Dict[str, Any]
    per_kp: List[Dict[str, Any]]
    ap_ar: Dict[str, Optional[float]]


def die(message: str, code: int = 2) -> None:
    print(f"ERROR: {message}", file=sys.stderr)
    raise SystemExit(code)


def load_json(path: Path) -> Any:
    try:
        with path.open("r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as exc:
        die(f"Cannot read JSON: {path}\n{exc}")


def resolve_path(raw: str, project_root: Optional[Path]) -> Path:
    if raw is None:
        die("Missing path value in config")
    s = os.path.expandvars(os.path.expanduser(str(raw).strip()))
    p = Path(s)
    if p.is_absolute() and p.exists():
        return p
    if project_root is not None:
        candidate = project_root / p
        if candidate.exists():
            return candidate
    if p.exists():
        return p
    # Handle JSON copied from Windows-like strings on a Linux host.
    normalized = Path(s.replace("\\", "/"))
    if normalized.is_absolute() and normalized.exists():
        return normalized
    if project_root is not None:
        candidate = project_root / normalized
        if candidate.exists():
            return candidate
    return p if project_root is None else (project_root / p)


def get_nested(d: Dict[str, Any], keys: Sequence[str], required: bool = True) -> Any:
    cur: Any = d
    for key in keys:
        if not isinstance(cur, dict) or key not in cur:
            if required:
                die(f"Missing config key: {'.'.join(keys)}")
            return None
        cur = cur[key]
    return cur


def normalize_keypoints(raw: Any) -> List[Tuple[float, float, float]]:
    """Return 17 tuples (x, y, score_or_visibility)."""
    if raw is None:
        return []
    if isinstance(raw, list) and len(raw) == 17 and all(isinstance(x, (list, tuple)) for x in raw):
        out = []
        for item in raw:
            if len(item) < 2:
                out.append((math.nan, math.nan, 0.0))
            else:
                score = float(item[2]) if len(item) > 2 and item[2] is not None else 1.0
                out.append((float(item[0]), float(item[1]), score))
        return out
    if isinstance(raw, list) and len(raw) >= 51:
        out = []
        for i in range(17):
            j = i * 3
            out.append((float(raw[j]), float(raw[j + 1]), float(raw[j + 2])))
        return out
    # Some YOLO exporters use a dict with xy/conf arrays.
    if isinstance(raw, dict):
        xy = raw.get("xy") or raw.get("xyn") or raw.get("points")
        conf = raw.get("conf") or raw.get("confidence") or raw.get("scores")
        if isinstance(xy, list) and len(xy) >= 17:
            out = []
            for i in range(17):
                pt = xy[i]
                score = 1.0
                if isinstance(conf, list) and i < len(conf):
                    score = float(conf[i])
                out.append((float(pt[0]), float(pt[1]), score))
            return out
    return []


def xyxy_to_xywh(b: Sequence[float]) -> List[float]:
    if len(b) >= 4:
        x1, y1, x2, y2 = [float(v) for v in b[:4]]
        return [x1, y1, max(0.0, x2 - x1), max(0.0, y2 - y1)]
    return []


def normalize_bbox(raw: Any) -> Optional[List[float]]:
    if raw is None:
        return None
    if isinstance(raw, dict):
        for key in ("bbox", "xywh", "bbox_xywh"):
            if key in raw:
                return normalize_bbox(raw[key])
        for key in ("bbox_xyxy", "xyxy"):
            if key in raw:
                return xyxy_to_xywh(raw[key])
    if isinstance(raw, list) and len(raw) >= 4:
        vals = [float(v) for v in raw[:4]]
        # Heuristic: if it looks like xyxy, convert; otherwise keep xywh.
        if vals[2] > vals[0] and vals[3] > vals[1] and (vals[2] - vals[0]) > 1 and (vals[3] - vals[1]) > 1:
            # Many exporters store xywh with width/height also greater than x/y for images near origin.
            # Prefer explicit key conversion when possible; keep generic lists as xywh.
            return vals
        return vals
    return None


def basename_no_query(path_or_name: str) -> str:
    value = str(path_or_name).replace("\\", "/")
    return os.path.basename(value.split("?")[0])


def load_gt(path: Path) -> Dict[str, Any]:
    raw = load_json(path)
    if not isinstance(raw, dict) or "images" not in raw or "annotations" not in raw:
        die(f"GT file is not COCO-style or lacks images/annotations: {path}")
    images_by_id = {int(img["id"]): img for img in raw.get("images", [])}
    image_id_by_name = {basename_no_query(img.get("file_name", "")): int(img["id"]) for img in raw.get("images", [])}
    anns_by_image: Dict[int, Dict[str, Any]] = {}
    for ann in raw.get("annotations", []):
        image_id = int(ann["image_id"])
        if image_id not in anns_by_image:
            anns_by_image[image_id] = ann
    return {
        "raw": raw,
        "path": str(path),
        "images_by_id": images_by_id,
        "image_id_by_name": image_id_by_name,
        "anns_by_image": anns_by_image,
    }


def parse_prediction_entry(entry: Dict[str, Any], image_id_by_name: Dict[str, int]) -> List[Prediction]:
    out: List[Prediction] = []
    # COCO/MMPose-style result: one dict == one prediction.
    if "keypoints" in entry and ("image_id" in entry or "image" in entry or "file_name" in entry):
        image_id: Optional[int]
        if "image_id" in entry:
            image_id = int(entry["image_id"])
        else:
            name = basename_no_query(entry.get("image") or entry.get("file_name") or "")
            image_id = image_id_by_name.get(name)
        if image_id is None:
            return out
        kps = normalize_keypoints(entry.get("keypoints"))
        if len(kps) == 17:
            score = float(entry.get("score", entry.get("bbox_score", np.mean([kp[2] for kp in kps]))))
            bbox = normalize_bbox(entry.get("bbox") or entry.get("bbox_xywh") or entry.get("bbox_xyxy"))
            out.append(Prediction(image_id=image_id, keypoints=kps, score=score, bbox=bbox))
        return out

    # YOLO image-level format: one dict per image with a list of predictions.
    preds = entry.get("predictions") or entry.get("detections") or entry.get("instances")
    if isinstance(preds, list):
        name = basename_no_query(entry.get("image") or entry.get("file_name") or entry.get("path") or "")
        image_id = int(entry["image_id"]) if "image_id" in entry else image_id_by_name.get(name)
        if image_id is None:
            return out
        for pred in preds:
            if not isinstance(pred, dict):
                continue
            kps = normalize_keypoints(pred.get("keypoints") or pred.get("kpts"))
            if len(kps) != 17:
                continue
            score = float(pred.get("score", pred.get("confidence", pred.get("conf", np.mean([kp[2] for kp in kps])))))
            bbox = normalize_bbox(pred.get("bbox") or pred.get("bbox_xywh") or pred.get("bbox_xyxy") or pred)
            out.append(Prediction(image_id=image_id, keypoints=kps, score=score, bbox=bbox))
    return out


def load_predictions(path: Path, image_id_by_name: Dict[str, int]) -> Dict[int, Prediction]:
    raw = load_json(path)
    if isinstance(raw, dict):
        if "annotations" in raw and isinstance(raw["annotations"], list):
            entries = raw["annotations"]
        elif "predictions" in raw and isinstance(raw["predictions"], list):
            entries = raw["predictions"]
        elif "results" in raw and isinstance(raw["results"], list):
            entries = raw["results"]
        else:
            entries = [raw]
    elif isinstance(raw, list):
        entries = raw
    else:
        die(f"Unsupported prediction JSON structure: {path}")

    grouped: Dict[int, List[Prediction]] = {}
    for entry in entries:
        if not isinstance(entry, dict):
            continue
        for pred in parse_prediction_entry(entry, image_id_by_name):
            grouped.setdefault(pred.image_id, []).append(pred)

    selected: Dict[int, Prediction] = {}
    for image_id, preds in grouped.items():
        selected[image_id] = max(preds, key=lambda p: p.score)
    return selected


def kp_accepted(pred: Optional[Prediction], kp_index: int, threshold: Optional[float]) -> bool:
    if pred is None or kp_index >= len(pred.keypoints):
        return False
    x, y, s = pred.keypoints[kp_index]
    if not math.isfinite(x) or not math.isfinite(y):
        return False
    if threshold is not None and s < threshold:
        return False
    return True


def annotation_area(ann: Dict[str, Any]) -> float:
    if "area" in ann and ann["area"]:
        return max(float(ann["area"]), 1.0)
    bbox = ann.get("bbox") or [0, 0, 1, 1]
    return max(float(bbox[2]) * float(bbox[3]), 1.0)


def bbox_diag(ann: Dict[str, Any]) -> float:
    bbox = ann.get("bbox") or [0, 0, 1, 1]
    return math.sqrt(float(bbox[2]) ** 2 + float(bbox[3]) ** 2)


def compute_error_rows(gt: Dict[str, Any], preds: Dict[int, Prediction], scenario: Scenario) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for image_id, ann in gt["anns_by_image"].items():
        pred = preds.get(image_id)
        gt_kps = normalize_keypoints(ann.get("keypoints"))
        diag = bbox_diag(ann)
        file_name = gt["images_by_id"].get(image_id, {}).get("file_name", "")
        for i, kp_name in enumerate(COCO_KEYPOINTS):
            if i >= len(gt_kps):
                continue
            gx, gy, gv = gt_kps[i]
            gt_visible = gv > 0
            accepted = kp_accepted(pred, i, scenario.threshold)
            px = py = ps = None
            err = None
            pck5 = False
            pck10 = False
            if pred is not None and i < len(pred.keypoints):
                px, py, ps = pred.keypoints[i]
            if gt_visible and accepted and px is not None and py is not None:
                err = math.sqrt((float(px) - float(gx)) ** 2 + (float(py) - float(gy)) ** 2)
                pck5 = (diag > 0 and err / diag <= 0.05)
                pck10 = (diag > 0 and err / diag <= 0.10)
            rows.append({
                "scenario_id": scenario.scenario_id,
                "dataset_train": scenario.dataset_train,
                "dataset_test": scenario.dataset_test,
                "model": scenario.model_label,
                "threshold": scenario.threshold_label,
                "image_id": image_id,
                "file_name": file_name,
                "kp_index": i,
                "kp": kp_name,
                "gt_visible": gt_visible,
                "pred_image_found": pred is not None,
                "pred_accepted": accepted,
                "gt_x": gx,
                "gt_y": gy,
                "pred_x": px,
                "pred_y": py,
                "pred_score": ps,
                "error_px": err,
                "pck5": pck5,
                "pck10": pck10,
                "invisible_gt_predicted": (not gt_visible and accepted),
            })
    return rows


def percentile(values: Sequence[float], q: float) -> Optional[float]:
    vals = [float(v) for v in values if v is not None and math.isfinite(float(v))]
    if not vals:
        return None
    return float(np.percentile(np.array(vals, dtype=np.float64), q))


def summarize_rows(rows: List[Dict[str, Any]], preds: Dict[int, Prediction], gt: Dict[str, Any], ap_ar: Optional[Dict[str, Optional[float]]] = None) -> Dict[str, Any]:
    visible_rows = [r for r in rows if r["gt_visible"]]
    invisible_rows = [r for r in rows if not r["gt_visible"]]
    valid_errors = [r["error_px"] for r in visible_rows if r["error_px"] is not None]
    visible_total = len(visible_rows)
    missing_visible = sum(1 for r in visible_rows if not r["pred_accepted"])
    invisible_pred = sum(1 for r in invisible_rows if r["invisible_gt_predicted"])
    invisible_total = len(invisible_rows)
    summary = {
        "images_total": len(gt["images_by_id"]),
        "images_with_predictions": len(set(preds.keys()) & set(gt["images_by_id"].keys())),
        "visible_kp_total": visible_total,
        "valid_visible_predictions": len(valid_errors),
        "missing_visible_kp": missing_visible,
        "missing_visible_rate": missing_visible / visible_total if visible_total else None,
        "invisible_kp_total": invisible_total,
        "invisible_gt_predicted": invisible_pred,
        "invisible_gt_predicted_rate": invisible_pred / invisible_total if invisible_total else None,
        "mean_error_px": float(mean(valid_errors)) if valid_errors else None,
        "median_error_px": float(median(valid_errors)) if valid_errors else None,
        "p90_error_px": percentile(valid_errors, 90),
        "p95_error_px": percentile(valid_errors, 95),
        "pck5": sum(1 for r in visible_rows if r["pck5"]) / visible_total if visible_total else None,
        "pck10": sum(1 for r in visible_rows if r["pck10"]) / visible_total if visible_total else None,
    }
    if ap_ar:
        summary.update(ap_ar)
    return summary


def per_kp_metrics(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for kp in COCO_KEYPOINTS:
        group = [r for r in rows if r["kp"] == kp]
        visible = [r for r in group if r["gt_visible"]]
        errors = [r["error_px"] for r in visible if r["error_px"] is not None]
        visible_total = len(visible)
        out.append({
            "kp": kp,
            "kp_index": COCO_KEYPOINTS.index(kp),
            "visible_kp_total": visible_total,
            "valid_visible_predictions": len(errors),
            "missing_visible_kp": sum(1 for r in visible if not r["pred_accepted"]),
            "missing_visible_rate": (sum(1 for r in visible if not r["pred_accepted"]) / visible_total) if visible_total else None,
            "mean_error_px": float(mean(errors)) if errors else None,
            "median_error_px": float(median(errors)) if errors else None,
            "p90_error_px": percentile(errors, 90),
            "pck5": sum(1 for r in visible if r["pck5"]) / visible_total if visible_total else None,
            "pck10": sum(1 for r in visible if r["pck10"]) / visible_total if visible_total else None,
        })
    return out


def make_coco_gt_filtered(gt_raw: Dict[str, Any], kp_indexes: Sequence[int]) -> Dict[str, Any]:
    keypoints_names = [COCO_KEYPOINTS[i] for i in kp_indexes]
    new = {
        "info": gt_raw.get("info", {}),
        "licenses": gt_raw.get("licenses", []),
        "images": gt_raw.get("images", []),
        "categories": [{
            "id": 1,
            "name": "person",
            "supercategory": "person",
            "keypoints": keypoints_names,
            "skeleton": [],
        }],
        "annotations": [],
    }
    for ann in gt_raw.get("annotations", []):
        old_kps = normalize_keypoints(ann.get("keypoints"))
        kps: List[float] = []
        visible_count = 0
        for idx in kp_indexes:
            x, y, v = old_kps[idx]
            kps.extend([x, y, v])
            if v > 0:
                visible_count += 1
        a = dict(ann)
        a["keypoints"] = kps
        a["num_keypoints"] = visible_count
        a["category_id"] = 1
        new["annotations"].append(a)
    return new


def make_coco_results(gt: Dict[str, Any], preds: Dict[int, Prediction], threshold: Optional[float], kp_indexes: Sequence[int]) -> List[Dict[str, Any]]:
    results: List[Dict[str, Any]] = []
    for image_id in gt["images_by_id"].keys():
        pred = preds.get(image_id)
        if pred is None:
            continue
        kps: List[float] = []
        accepted_scores = []
        for idx in kp_indexes:
            if kp_accepted(pred, idx, threshold):
                x, y, s = pred.keypoints[idx]
                kps.extend([float(x), float(y), float(s)])
                accepted_scores.append(float(s))
            else:
                # COCOeval will count this as a large miss for visible GT keypoints.
                kps.extend([0.0, 0.0, 0.0])
        score = float(pred.score)
        if not math.isfinite(score):
            score = float(np.mean(accepted_scores)) if accepted_scores else 0.0
        results.append({
            "image_id": int(image_id),
            "category_id": 1,
            "keypoints": kps,
            "score": score,
        })
    return results


def compute_ap_ar(gt: Dict[str, Any], preds: Dict[int, Prediction], threshold: Optional[float], kp_indexes: Sequence[int] = ALL_KP_INDEXES) -> Dict[str, Optional[float]]:
    if COCO is not None and COCOeval is not None:
        try:
            with tempfile.TemporaryDirectory() as td:
                gt_file = Path(td) / "gt.json"
                res_file = Path(td) / "res.json"
                gt_filtered = make_coco_gt_filtered(gt["raw"], kp_indexes)
                results = make_coco_results(gt, preds, threshold, kp_indexes)
                gt_file.write_text(json.dumps(gt_filtered), encoding="utf-8")
                res_file.write_text(json.dumps(results), encoding="utf-8")
                with contextlib.redirect_stdout(io.StringIO()):
                    coco_gt = COCO(str(gt_file))
                    coco_dt = coco_gt.loadRes(str(res_file)) if results else coco_gt.loadRes([])
                    coco_eval = COCOeval(coco_gt, coco_dt, "keypoints")
                    coco_eval.params.imgIds = list(gt["images_by_id"].keys())
                    coco_eval.params.kpt_oks_sigmas = COCO_SIGMAS[list(kp_indexes)]
                    coco_eval.evaluate()
                    coco_eval.accumulate()
                    coco_eval.summarize()
                stats = list(coco_eval.stats)
                return {
                    "AP": float(stats[0]),
                    "AP50": float(stats[1]),
                    "AP75": float(stats[2]),
                    "AR": float(stats[5]),
                    "AR50": float(stats[6]),
                    "AR75": float(stats[7]),
                    "ap_ar_method": "pycocotools",
                }
        except Exception as exc:
            print(f"WARNING: pycocotools AP/AR failed; using fallback approximate OKS metrics. Reason: {exc}", file=sys.stderr)
    return compute_ap_ar_fallback(gt, preds, threshold, kp_indexes)


def compute_ap_ar_fallback(gt: Dict[str, Any], preds: Dict[int, Prediction], threshold: Optional[float], kp_indexes: Sequence[int]) -> Dict[str, Optional[float]]:
    oks_values: List[float] = []
    sigmas = COCO_SIGMAS[list(kp_indexes)]
    vars_ = (sigmas * 2) ** 2
    for image_id, ann in gt["anns_by_image"].items():
        gt_kps = normalize_keypoints(ann.get("keypoints"))
        pred = preds.get(image_id)
        if pred is None:
            oks_values.append(0.0)
            continue
        area = annotation_area(ann)
        vals = []
        for pos, idx in enumerate(kp_indexes):
            gx, gy, gv = gt_kps[idx]
            if gv <= 0:
                continue
            if not kp_accepted(pred, idx, threshold):
                vals.append(0.0)
                continue
            px, py, _ = pred.keypoints[idx]
            d2 = (px - gx) ** 2 + (py - gy) ** 2
            vals.append(float(math.exp(-d2 / (vars_[pos] * (area + np.spacing(1)) * 2))))
        oks_values.append(float(mean(vals)) if vals else 0.0)
    recalls = [sum(1 for oks in oks_values if oks >= t) / len(oks_values) for t in OKS_THRESHOLDS] if oks_values else []
    m = float(mean(recalls)) if recalls else None
    return {
        "AP": m,
        "AP50": recalls[0] if recalls else None,
        "AP75": recalls[5] if len(recalls) > 5 else None,
        "AR": m,
        "AR50": recalls[0] if recalls else None,
        "AR75": recalls[5] if len(recalls) > 5 else None,
        "ap_ar_method": "fallback_single_prediction_oks",
    }


def safe_float(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    try:
        return float(str(value).strip())
    except ValueError:
        return None


def find_col(headers: Sequence[str], candidates: Sequence[str]) -> Optional[str]:
    normalized = {h.lower().strip(): h for h in headers}
    for cand in candidates:
        if cand.lower().strip() in normalized:
            return normalized[cand.lower().strip()]
    return None


def extract_val_best(csv_path: Path, model_key: str, min_delta: float, patience: int) -> Dict[str, Any]:
    with csv_path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        headers = reader.fieldnames or []
    if not rows:
        die(f"Validation CSV has no rows: {csv_path}")
    ap_col = find_col(headers, ["AP", "metrics/mAP50-95(P)", "metrics/mAP50-95(B)", "mAP50-95", "map"])
    ar_col = find_col(headers, ["AR", "metrics/recall(P)", "metrics/recall(B)", "recall", "AR50-95"])
    epoch_col = find_col(headers, ["epoch", "Epoch"])
    if ap_col is None or ar_col is None:
        die(f"Cannot identify AP/AR columns in validation CSV: {csv_path}. Headers={headers}")
    best_row = None
    best_ap = -float("inf")
    bad_count = 0
    stop_epoch = None
    for row in rows:
        ap = safe_float(row.get(ap_col))
        if ap is None:
            continue
        if best_row is None or ap > best_ap + min_delta:
            best_row = row
            best_ap = ap
            bad_count = 0
        else:
            bad_count += 1
        if bad_count >= patience:
            stop_epoch = row.get(epoch_col) if epoch_col else None
            break
    if best_row is None:
        die(f"No valid AP values in validation CSV: {csv_path}")
    return {
        "csv_path": str(csv_path),
        "model_key": model_key,
        "epoch": int(float(best_row.get(epoch_col, 0))) if epoch_col and best_row.get(epoch_col) else None,
        "AP": safe_float(best_row.get(ap_col)),
        "AR": safe_float(best_row.get(ar_col)),
        "ap_col": ap_col,
        "ar_col": ar_col,
        "min_delta": min_delta,
        "patience": patience,
        "stop_epoch": stop_epoch,
    }


def fmt_num(value: Optional[float], digits: int = 3) -> str:
    if value is None:
        return "n/a"
    return f"{value:.{digits}f}"


def fmt_pct(value: Optional[float], digits: int = 3) -> str:
    if value is None:
        return "n/a"
    return f"{value * 100:.{digits}f}%"


def best_label(yolo: Optional[float], vit: Optional[float], higher_better: bool) -> str:
    if yolo is None or vit is None:
        return "n/a"
    diff = yolo - vit
    if abs(diff) < 1e-4:
        return "≈"
    if higher_better:
        return "YOLO" if yolo > vit else "VitPose++"
    return "YOLO" if yolo < vit else "VitPose++"


def trend(value: Optional[float], baseline: Optional[float], eps: float) -> str:
    if value is None or baseline is None:
        return ""
    if value > baseline + eps:
        return "↑"
    if value < baseline - eps:
        return "↓"
    return "↔"


def get_bundle(bundles: Dict[str, EvalBundle], scenario_id: str, model_key: str) -> EvalBundle:
    key = f"{scenario_id}::{model_key}"
    if key not in bundles:
        die(f"Internal error: missing bundle {key}")
    return bundles[key]


def get_per_kp(bundle: EvalBundle, kp: str) -> Dict[str, Any]:
    for row in bundle.per_kp:
        if row["kp"] == kp:
            return row
    return {}


def combined_p90_by_kp(yolo_bundle: EvalBundle, vit_bundle: EvalBundle) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for kp in COCO_KEYPOINTS:
        errors = []
        for bundle in (yolo_bundle, vit_bundle):
            errors.extend([r["error_px"] for r in bundle.error_rows if r["kp"] == kp and r["error_px"] is not None])
        out.append({"kp": kp, "p90_combined": percentile(errors, 90)})
    out.sort(key=lambda r: (float("inf") if r["p90_combined"] is None else r["p90_combined"]))
    return out


def classify_p90(value: Optional[float], thresholds: Dict[str, float]) -> str:
    if value is None:
        return "n/a"
    if value <= thresholds["easy_max"]:
        return "Easy"
    if value <= thresholds["medium_max"]:
        return "Medium"
    if value <= thresholds["high_max"]:
        return "High"
    return "Challenging"


def make_scenarios(config: Dict[str, Any], yolo_thr: float, vit_thr: float) -> List[Scenario]:
    model_info = {
        "yolo": ("YOLO26x-Pose", yolo_thr),
        "vitpose": ("VitPose++", vit_thr),
    }
    scenarios: List[Scenario] = []
    for base_id, train, test, gt_key, pred_prefix in [
        ("direct_A", "SAW_frames_EntireSwim", "SAW_frames_EntireSwim", "saw_frames_entireswim", "direct.saw_frames_entireswim"),
        ("direct_B", "SAW_frames", "SAW_frames", "saw_frames", "direct.saw_frames"),
        ("cross_B_to_A", "SAW_frames", "SAW_frames_EntireSwim", "saw_frames_entireswim", "cross.train_saw_frames_test_saw_frames_entireswim"),
    ]:
        for model_key, (model_label, thr) in model_info.items():
            for threshold_label, threshold in [("T0", None), ("Top", thr)]:
                pred_key = f"{pred_prefix}.{model_key}"
                scenarios.append(Scenario(
                    scenario_id=f"{base_id}_{threshold_label}",
                    dataset_train=train,
                    dataset_test=test,
                    gt_key=gt_key,
                    model_key=model_key,
                    model_label=model_label,
                    pred_key=pred_key,
                    threshold=threshold,
                    threshold_label=("0" if threshold is None else str(threshold)),
                ))
    return scenarios


def flatten_paths(paths: Dict[str, Any]) -> Dict[str, str]:
    out: Dict[str, str] = {}
    def rec(prefix: str, obj: Any) -> None:
        if isinstance(obj, dict):
            for k, v in obj.items():
                rec(f"{prefix}.{k}" if prefix else k, v)
        else:
            out[prefix] = str(obj)
    rec("", paths)
    return out


def evaluate_all(config: Dict[str, Any], args: argparse.Namespace) -> Tuple[Dict[str, EvalBundle], Dict[str, Dict[str, Any]]]:
    project_root = resolve_path(config.get("project_root", "."), None) if config.get("project_root") else None
    paths = get_nested(config, ["paths"])
    flat = flatten_paths(paths)

    gt_paths = {
        "saw_frames": resolve_path(get_nested(paths, ["gt", "saw_frames"]), project_root),
        "saw_frames_entireswim": resolve_path(get_nested(paths, ["gt", "saw_frames_entireswim"]), project_root),
    }
    gt_cache = {key: load_gt(path) for key, path in gt_paths.items()}

    bundles: Dict[str, EvalBundle] = {}
    pred_cache: Dict[Tuple[str, str], Dict[int, Prediction]] = {}
    scenarios = make_scenarios(config, args.yolo_threshold, args.vitpose_threshold)
    for sc in scenarios:
        gt = gt_cache[sc.gt_key]
        raw_pred_path = flat.get(f"predictions.{sc.pred_key}")
        if raw_pred_path is None:
            die(f"Missing prediction path in config: paths.predictions.{sc.pred_key}")
        pred_path = resolve_path(raw_pred_path, project_root)
        if not pred_path.exists():
            die(f"Prediction file does not exist: {pred_path}")
        cache_key = (str(pred_path), sc.gt_key)
        if cache_key not in pred_cache:
            pred_cache[cache_key] = load_predictions(pred_path, gt["image_id_by_name"])
        preds = pred_cache[cache_key]
        rows = compute_error_rows(gt, preds, sc)
        ap_ar = compute_ap_ar(gt, preds, sc.threshold, ALL_KP_INDEXES)
        summary = summarize_rows(rows, preds, gt, ap_ar)
        per_kp = per_kp_metrics(rows)
        bundles[f"{sc.scenario_id}::{sc.model_key}"] = EvalBundle(
            scenario=sc,
            gt_path=gt_paths[sc.gt_key],
            pred_path=pred_path,
            gt=gt,
            predictions=preds,
            error_rows=rows,
            summary=summary,
            per_kp=per_kp,
            ap_ar=ap_ar,
        )

    val_metrics: Dict[str, Dict[str, Any]] = {}
    val_paths = get_nested(paths, ["val_metrics"], required=False) or {}
    for dataset_key in ("saw_frames", "saw_frames_entireswim"):
        for model_key in ("yolo", "vitpose"):
            raw = get_nested(val_paths, [dataset_key, model_key], required=False)
            if raw:
                p = resolve_path(raw, project_root)
                if not p.exists():
                    die(f"Validation CSV does not exist: {p}")
                val_metrics[f"{dataset_key}::{model_key}"] = extract_val_best(p, model_key, args.min_delta, args.patience)
    return bundles, val_metrics


# ----------------------------- XLSX helpers -----------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
TITLE_FILL = PatternFill("solid", fgColor="0B1F33")
WHITE_FONT = Font(color="FFFFFF", bold=True)
BOLD_FONT = Font(bold=True)
SMALL_FONT = Font(size=9)
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def new_sheet(wb: Workbook, name: str, title: str) -> Any:
    ws = wb.create_sheet(title=name[:31])
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"
    ws["A1"] = title
    ws["A1"].font = Font(size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 24
    return ws


def write_table(ws: Any, start_row: int, start_col: int, title: str, headers: List[str], rows: List[List[Any]], note: Optional[str] = None) -> Tuple[int, int]:
    c0 = start_col
    r = start_row
    ws.cell(r, c0, title)
    ws.cell(r, c0).font = Font(bold=True, color="FFFFFF")
    ws.cell(r, c0).fill = HEADER_FILL
    ws.cell(r, c0).alignment = Alignment(horizontal="left")
    if len(headers) > 1:
        ws.merge_cells(start_row=r, start_column=c0, end_row=r, end_column=c0 + len(headers) - 1)
    r += 1
    for j, h in enumerate(headers):
        cell = ws.cell(r, c0 + j, h)
        cell.font = WHITE_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    for row in rows:
        r += 1
        for j, val in enumerate(row):
            cell = ws.cell(r, c0 + j, val)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center" if isinstance(val, (int, float)) else "left", vertical="center", wrap_text=True)
            if isinstance(val, float):
                cell.number_format = "0.000"
    end_row = r
    if note:
        r += 2
        ws.cell(r, c0, note)
        ws.cell(r, c0).font = SMALL_FONT
        ws.cell(r, c0).alignment = Alignment(wrap_text=True)
        if len(headers) > 1:
            ws.merge_cells(start_row=r, start_column=c0, end_row=r, end_column=c0 + len(headers) - 1)
        end_row = r
    return end_row, c0 + len(headers) - 1


def autosize(ws: Any) -> None:
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 8
        for cell in ws[letter]:
            if cell.value is None:
                continue
            max_len = max(max_len, min(42, len(str(cell.value)) + 2))
        ws.column_dimensions[letter].width = max_len
    for row_idx in range(1, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = max(ws.row_dimensions[row_idx].height or 15, 18)


def add_conditional_arrows(ws: Any) -> None:
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                if "↑" in cell.value:
                    cell.font = Font(color="008000", bold=cell.font.bold)
                elif "↓" in cell.value:
                    cell.font = Font(color="C00000", bold=cell.font.bold)
                elif "↔" in cell.value:
                    cell.font = Font(color="666666", bold=cell.font.bold)


def style_workbook(wb: Workbook) -> None:
    for ws in wb.worksheets:
        autosize(ws)
        add_conditional_arrows(ws)


def slide8(wb: Workbook, bundles: Dict[str, EvalBundle], val_metrics: Dict[str, Dict[str, Any]]) -> None:
    ws = new_sheet(wb, "Slide 8", "Slide 8 - Basic Test Overall HPE")
    def val(dataset: str, model: str, metric: str) -> Any:
        return val_metrics.get(f"{dataset}::{model}", {}).get(metric)
    rows = []
    for model_key, label in [("yolo", "YOLO26x-Pose"), ("vitpose", "YOLO26x-Det → VitPose++")]:
        direct_a = get_bundle(bundles, "direct_A_Top", model_key).summary
        direct_b = get_bundle(bundles, "direct_B_Top", model_key).summary
        rows.append([
            label,
            val("saw_frames_entireswim", model_key, "AR"),
            val("saw_frames_entireswim", model_key, "AP"),
            direct_a.get("AR"), direct_a.get("AP"),
            val("saw_frames", model_key, "AR"),
            val("saw_frames", model_key, "AP"),
            direct_b.get("AR"), direct_b.get("AP"),
        ])
    write_table(ws, 3, 1, "SAW_frames_EntireSwim and SAW_frames - Val/Test AP/AR", [
        "Model", "EntireSwim Val AR", "EntireSwim Val AP", "EntireSwim Test AR", "EntireSwim Test AP",
        "SAW_frames Val AR", "SAW_frames Val AP", "SAW_frames Test AR", "SAW_frames Test AP",
    ], rows, "Val values are extracted from training CSV with patience/min_delta; Test values are recomputed from GT and KP JSON.")


def slide9(wb: Workbook, bundles: Dict[str, EvalBundle]) -> None:
    ws = new_sheet(wb, "Slide 9", "Slide 9 - Basic Test Overall Geometric Accuracy")
    for start_row, scenario, title in [(3, "direct_A_Top", "Dataset SAW_frames_EntireSwim"), (13, "direct_B_Top", "Dataset SAW_frames")]:
        y = get_bundle(bundles, scenario, "yolo").summary
        v = get_bundle(bundles, scenario, "vitpose").summary
        rows = [
            ["Mean Error (px)", y.get("mean_error_px"), v.get("mean_error_px"), best_label(y.get("mean_error_px"), v.get("mean_error_px"), False)],
            ["Median Error (px)", y.get("median_error_px"), v.get("median_error_px"), best_label(y.get("median_error_px"), v.get("median_error_px"), False)],
            ["P90 Error (px)", y.get("p90_error_px"), v.get("p90_error_px"), best_label(y.get("p90_error_px"), v.get("p90_error_px"), False)],
            ["PCK@5", y.get("pck5"), v.get("pck5"), best_label(y.get("pck5"), v.get("pck5"), True)],
            ["PCK@10", y.get("pck10"), v.get("pck10"), best_label(y.get("pck10"), v.get("pck10"), True)],
        ]
        write_table(ws, start_row, 1, title, ["Metric", "YOLO26x-Pose", "VitPose++", "Best"], rows)


def slide10(wb: Workbook, bundles: Dict[str, EvalBundle]) -> None:
    ws = new_sheet(wb, "Slide 10", "Slide 10 - Basic Test Overall Reliability")
    for start_row, scenario, title in [(3, "direct_A_Top", "Dataset SAW_frames_EntireSwim"), (11, "direct_B_Top", "Dataset SAW_frames")]:
        y = get_bundle(bundles, scenario, "yolo").summary
        v = get_bundle(bundles, scenario, "vitpose").summary
        rows = [
            ["Frames with predictions", f"{y['images_with_predictions']} / {y['images_total']}", f"{v['images_with_predictions']} / {v['images_total']}", best_label(y.get("images_with_predictions"), v.get("images_with_predictions"), True)],
            ["Missing Visible KP", y.get("missing_visible_rate"), v.get("missing_visible_rate"), best_label(y.get("missing_visible_rate"), v.get("missing_visible_rate"), False)],
            ["Invisible KP from GT", y.get("invisible_gt_predicted_rate"), v.get("invisible_gt_predicted_rate"), best_label(y.get("invisible_gt_predicted_rate"), v.get("invisible_gt_predicted_rate"), False)],
        ]
        write_table(ws, start_row, 1, title, ["Metric", "YOLO26x-Pose", "VitPose++", "Best"], rows)


def slide16(wb: Workbook, bundles: Dict[str, EvalBundle]) -> None:
    ws = new_sheet(wb, "Slide 16", "Slide 16 - Cross Test Overall HPE")
    rows = []
    for model_key, label in [("yolo", "YOLO26x-Pose"), ("vitpose", "VitPose++")]:
        direct_top = get_bundle(bundles, "direct_A_Top", model_key).summary
        cross_top = get_bundle(bundles, "cross_B_to_A_Top", model_key).summary
        direct_t0 = get_bundle(bundles, "direct_A_T0", model_key).summary
        cross_t0 = get_bundle(bundles, "cross_B_to_A_T0", model_key).summary
        rows.append([
            label,
            f"{fmt_num(cross_top.get('AR'), 4)} {trend(cross_top.get('AR'), direct_top.get('AR'), 0.0005)}",
            f"{fmt_num(cross_top.get('AP'), 4)} {trend(cross_top.get('AP'), direct_top.get('AP'), 0.0005)}",
            f"{fmt_num(cross_t0.get('AR'), 4)} {trend(cross_t0.get('AR'), direct_t0.get('AR'), 0.0005)}",
            f"{fmt_num(cross_t0.get('AP'), 4)} {trend(cross_t0.get('AP'), direct_t0.get('AP'), 0.0005)}",
        ])
    write_table(ws, 3, 1, "Train SAW_frames → Test SAW_frames_EntireSwim", [
        "Model", "With Threshold AR", "With Threshold AP", "Without Threshold AR", "Without Threshold AP",
    ], rows, "Arrows compare cross-test against direct Train SAW_frames_EntireSwim → Test SAW_frames_EntireSwim.")


def slide17(wb: Workbook, bundles: Dict[str, EvalBundle]) -> None:
    ws = new_sheet(wb, "Slide 17", "Slide 17 - Cross Test Single KPs")
    y = get_bundle(bundles, "cross_B_to_A_Top", "yolo")
    v = get_bundle(bundles, "cross_B_to_A_Top", "vitpose")
    rows = []
    for kp in COCO_KEYPOINTS:
        yy = get_per_kp(y, kp)
        vv = get_per_kp(v, kp)
        rows.append([
            kp,
            f"{fmt_num(yy.get('mean_error_px'), 2)} / {fmt_num(yy.get('median_error_px'), 2)} / {fmt_num(yy.get('p90_error_px'), 2)}",
            f"{fmt_num(vv.get('mean_error_px'), 2)} / {fmt_num(vv.get('median_error_px'), 2)} / {fmt_num(vv.get('p90_error_px'), 2)}",
        ])
    write_table(ws, 3, 1, "Train SAW_frames → Test SAW_frames_EntireSwim (thresholded)", ["KP", "YOLO26x-Pose mean / median / P90", "VitPose++ mean / median / P90"], rows)


def direct_vs_cross_rows(bundles: Dict[str, EvalBundle], model_key: str) -> List[List[Any]]:
    d = get_bundle(bundles, "direct_A_Top", model_key)
    c = get_bundle(bundles, "cross_B_to_A_Top", model_key)
    rows = []
    for kp in COCO_KEYPOINTS:
        dd = get_per_kp(d, kp)
        cc = get_per_kp(c, kp)
        dm = (cc.get("mean_error_px") - dd.get("mean_error_px")) if dd.get("mean_error_px") is not None and cc.get("mean_error_px") is not None else None
        dmed = (cc.get("median_error_px") - dd.get("median_error_px")) if dd.get("median_error_px") is not None and cc.get("median_error_px") is not None else None
        dp90 = (cc.get("p90_error_px") - dd.get("p90_error_px")) if dd.get("p90_error_px") is not None and cc.get("p90_error_px") is not None else None
        rows.append([kp, f"{fmt_num(dd.get('mean_error_px'), 3)} → {fmt_num(cc.get('mean_error_px'), 3)}", dm, dmed, dp90])
    rows.sort(key=lambda r: float("inf") if r[2] is None else r[2])
    return rows


def slide18(wb: Workbook, bundles: Dict[str, EvalBundle]) -> None:
    ws = new_sheet(wb, "Slide 18", "Slide 18 - Cross Test Direct vs Cross Single KPs")
    headers = ["KP", "Mean direct → cross", "Δ mean", "Δ median", "Δ P90"]
    write_table(ws, 3, 1, "YOLO26x-Pose", headers, direct_vs_cross_rows(bundles, "yolo"))
    write_table(ws, 3, 8, "VitPose++", headers, direct_vs_cross_rows(bundles, "vitpose"), "Negative Δ means improvement from Train A to Train B on the same Test A.")


def slide19_and_20_data(bundles: Dict[str, EvalBundle], thresholds: Dict[str, float]) -> Tuple[List[List[Any]], List[List[Any]]]:
    y = get_bundle(bundles, "cross_B_to_A_Top", "yolo")
    v = get_bundle(bundles, "cross_B_to_A_Top", "vitpose")
    combined = combined_p90_by_kp(y, v)
    rows19: List[List[Any]] = []
    rows20: List[List[Any]] = []
    for rank, item in enumerate(combined, start=1):
        kp = item["kp"]
        p90 = item["p90_combined"]
        group = classify_p90(p90, thresholds)
        label = f"{group} ●" if kp in HEAD_KP else group
        rows19.append([rank, kp, p90, label])
        yy = get_per_kp(y, kp)
        vv = get_per_kp(v, kp)
        ymean = yy.get("mean_error_px")
        vmean = vv.get("mean_error_px")
        delta = (ymean - vmean) if ymean is not None and vmean is not None else None
        gain = (delta / ymean) if ymean not in (None, 0) and delta is not None else None
        rows20.append([kp, label, p90, ymean, vmean, delta, gain])
    return rows19, rows20


def slide19(wb: Workbook, bundles: Dict[str, EvalBundle], thresholds: Dict[str, float]) -> None:
    ws = new_sheet(wb, "Slide 19", "Slide 19 - Cross Test KP Difficulty")
    rows19, _ = slide19_and_20_data(bundles, thresholds)
    write_table(ws, 3, 1, "Comparison of KP detection difficulty", ["Rank", "KP", "Combined distributions over models P90 px ↑", "Classification"], rows19, "● indicates head KPs. P90 is computed from the unified YOLO+VitPose error distribution for each KP.")


def slide20(wb: Workbook, bundles: Dict[str, EvalBundle], thresholds: Dict[str, float]) -> None:
    ws = new_sheet(wb, "Slide 20", "Slide 20 - Cross Test KP Detection Capability")
    _, rows20 = slide19_and_20_data(bundles, thresholds)
    write_table(ws, 3, 1, "Analysis of the single KPs detection capability of the models", [
        "KP", "Group", "P90 comb.", "YOLO mean err", "VitPose mean err", "Δ mean err", "Gain VitPose",
    ], rows20, "Δ mean = YOLO mean − VitPose mean. Positive values indicate VitPose++ advantage.")


def build_workbook(bundles: Dict[str, EvalBundle], val_metrics: Dict[str, Dict[str, Any]], output_path: Path, difficulty_thresholds: Dict[str, float], slides: Sequence[int]) -> None:
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    if 8 in slides:
        slide8(wb, bundles, val_metrics)
    if 9 in slides:
        slide9(wb, bundles)
    if 10 in slides:
        slide10(wb, bundles)
    if 16 in slides:
        slide16(wb, bundles)
    if 17 in slides:
        slide17(wb, bundles)
    if 18 in slides:
        slide18(wb, bundles)
    if 19 in slides:
        slide19(wb, bundles, difficulty_thresholds)
    if 20 in slides:
        slide20(wb, bundles, difficulty_thresholds)
    style_workbook(wb)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def write_csv(path: Path, rows: List[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        return
    headers = list(rows[0].keys())
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def export_intermediates(out_dir: Path, bundles: Dict[str, EvalBundle], val_metrics: Dict[str, Dict[str, Any]]) -> None:
    summaries = []
    per_kp_all = []
    for key, bundle in bundles.items():
        row = dict(bundle.summary)
        row.update({
            "scenario_id": bundle.scenario.scenario_id,
            "model_key": bundle.scenario.model_key,
            "model": bundle.scenario.model_label,
            "gt_path": str(bundle.gt_path),
            "pred_path": str(bundle.pred_path),
            "threshold": bundle.scenario.threshold_label,
        })
        summaries.append(row)
        for kp_row in bundle.per_kp:
            r = dict(kp_row)
            r.update({
                "scenario_id": bundle.scenario.scenario_id,
                "model_key": bundle.scenario.model_key,
                "model": bundle.scenario.model_label,
                "threshold": bundle.scenario.threshold_label,
            })
            per_kp_all.append(r)
    write_csv(out_dir / "scenario_summary.csv", summaries)
    write_csv(out_dir / "per_kp_metrics.csv", per_kp_all)
    write_csv(out_dir / "val_best_epoch_summary.csv", list(val_metrics.values()))


def parse_slides(raw: str) -> List[int]:
    if raw.lower() in ("all", "*"):
        return [8, 9, 10, 16, 17, 18, 19, 20]
    values = []
    for part in raw.split(","):
        part = part.strip()
        if not part:
            continue
        values.append(int(part))
    return values


def main() -> None:
    parser = argparse.ArgumentParser(description="Build HPE report tables XLSX from GT/prediction JSON and Val CSV metrics.")
    parser.add_argument("--config", required=True, help="Path to JSON config file.")
    parser.add_argument("--output", default="data/output/experiments/hpe_report/hpe_report_tables.xlsx", help="Output XLSX path.")
    parser.add_argument("--yolo-threshold", type=float, default=0.30, help="Default YOLO keypoint confidence threshold for thresholded tables.")
    parser.add_argument("--vitpose-threshold", type=float, default=0.20, help="Default VitPose keypoint confidence threshold for thresholded tables.")
    parser.add_argument("--min-delta", type=float, default=0.007, help="Min delta used to reproduce validation best epoch selection.")
    parser.add_argument("--patience", type=int, default=3, help="Patience used to reproduce validation best epoch selection.")
    parser.add_argument("--slides", default="all", help="Comma-separated slide numbers to export, e.g. 8,16,17,18,19,20 or all.")
    parser.add_argument("--validate-only", action="store_true", help="Compute metrics and validate inputs without writing XLSX.")
    parser.add_argument("--export-intermediate-csv", action="store_true", help="Export scenario_summary/per_kp_metrics/val CSV files next to the XLSX.")
    parser.add_argument("--difficulty-easy-max", type=float, default=6.0, help="P90 max threshold for Easy KP group.")
    parser.add_argument("--difficulty-medium-max", type=float, default=9.0, help="P90 max threshold for Medium KP group.")
    parser.add_argument("--difficulty-high-max", type=float, default=12.0, help="P90 max threshold for High KP group; above is Challenging.")
    args = parser.parse_args()

    config_path = Path(args.config)
    if not config_path.exists():
        die(f"Config file does not exist: {config_path}")
    config = load_json(config_path)
    if not isinstance(config, dict):
        die("Config root must be a JSON object.")

    output_path = resolve_path(args.output, resolve_path(config.get("project_root", "."), None) if config.get("project_root") else None)
    bundles, val_metrics = evaluate_all(config, args)

    print(f"Validated/evaluated {len(bundles)} model-scenarios.")
    methods = sorted(set(str(b.ap_ar.get("ap_ar_method")) for b in bundles.values()))
    print(f"AP/AR method(s): {', '.join(methods)}")

    if args.export_intermediate_csv:
        export_intermediates(output_path.parent / "intermediate_csv", bundles, val_metrics)
        print(f"Intermediate CSV exported to: {output_path.parent / 'intermediate_csv'}")

    if args.validate_only:
        print("Validation only completed; XLSX not written.")
        return

    difficulty_thresholds = {
        "easy_max": args.difficulty_easy_max,
        "medium_max": args.difficulty_medium_max,
        "high_max": args.difficulty_high_max,
    }
    build_workbook(bundles, val_metrics, output_path, difficulty_thresholds, parse_slides(args.slides))
    print(f"XLSX written to: {output_path}")


if __name__ == "__main__":
    main()
